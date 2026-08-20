"""
API endpoints for Import/Export transactions (Nhập/Xuất kho).
"""
import math
from fastapi import APIRouter, Depends, HTTPException, Query
from sqlalchemy.orm import Session
from sqlalchemy import or_
from typing import Optional
from datetime import date
from sqlalchemy import or_, func

from ..deps import require_permission, apply_warehouse_filter, check_warehouse_access
from ..database import get_db
from ..models import Item, Transaction
from ..schemas import (
    ImportCreate, ExportCreate,
    TransactionResponse, TransactionListResponse,
)

router = APIRouter(prefix="/api/transactions", tags=["Nhập/Xuất kho"])


@router.get("", response_model=TransactionListResponse)
def list_transactions(
    loai: Optional[str] = None,  # NHAP or XUAT
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
    search: Optional[str] = None,
    from_date: Optional[date] = None,
    to_date: Optional[date] = None,
    cong_doan: Optional[str] = None,
    nguoi: Optional[str] = None,
    kho_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user = Depends(require_permission("perm_view"))
):
    """Lấy danh sách giao dịch nhập/xuất kho."""
    query = db.query(Transaction)
    
    query = apply_warehouse_filter(query, Transaction, current_user, db, kho_id)

    # Type filter
    if loai:
        query = query.filter(Transaction.loai == loai)

    # Search
    if search:
        search_lower = search.lower()
        query = query.filter(
            or_(
                func.lower(Transaction.ten_hang).contains(search_lower),
                func.lower(Transaction.ma_so).contains(search_lower),
            )
        )

    # Date range filter
    if from_date:
        query = query.filter(Transaction.ngay >= from_date)
    if to_date:
        query = query.filter(Transaction.ngay <= to_date)

    # Filters
    if cong_doan:
        query = query.filter(func.lower(Transaction.cong_doan).contains(cong_doan.lower()))
    if nguoi:
        nguoi_lower = nguoi.lower()
        query = query.filter(
            or_(
                func.lower(Transaction.nguoi_nhap).contains(nguoi_lower),
                func.lower(Transaction.nguoi_yeu_cau).contains(nguoi_lower),
                func.lower(Transaction.nguoi_nhan).contains(nguoi_lower),
                func.lower(Transaction.nguoi_xuat).contains(nguoi_lower),
            )
        )

    # Count
    total = query.count()

    # Sort by date descending (newest first)
    query = query.order_by(Transaction.ngay.desc(), Transaction.id.desc())

    # Paginate
    offset = (page - 1) * page_size
    transactions = query.offset(offset).limit(page_size).all()

    return TransactionListResponse(
        transactions=[TransactionResponse.model_validate(t) for t in transactions],
        total=total,
        page=page,
        page_size=page_size,
        total_pages=math.ceil(total / page_size) if total > 0 else 1,
    )


@router.post("/import", response_model=TransactionResponse, status_code=201)
def create_import(data: ImportCreate, db: Session = Depends(get_db), current_user = Depends(require_permission("perm_add"))):
    """Tạo phiếu nhập kho - tự động cộng tồn kho."""
    check_warehouse_access(current_user, data.kho_id)
    # Verify item exists
    item = db.query(Item).filter(Item.id == data.item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Không tìm thấy mặt hàng")
    if item.kho_id != data.kho_id:
        raise HTTPException(status_code=400, detail="Mặt hàng không thuộc kho này")

    # Create transaction
    transaction = Transaction(
        loai="NHAP",
        item_id=item.id,
        ngay=data.ngay,
        ma_quan_ly=item.ma_quan_ly,
        ten_hang=item.ten_hang,
        ma_so=item.ma_so,
        so_luong=data.so_luong,
        don_vi_tinh=item.don_vi_tinh,
        cong_doan=data.cong_doan,
        nguoi_nhap=data.nguoi_nhap,
        trang_thai=data.trang_thai,
        ghi_chu=data.ghi_chu,
        kho_id=data.kho_id or item.kho_id
    )
    db.add(transaction)

    # Update item stock
    item.tong_nhap += data.so_luong
    item.ton_cuoi = item.ton_dau + item.tong_nhap - item.tong_xuat

    db.commit()
    db.refresh(transaction)
    return transaction


@router.post("/export", response_model=TransactionResponse, status_code=201)
def create_export(data: ExportCreate, db: Session = Depends(get_db), current_user = Depends(require_permission("perm_add"))):
    """Tạo phiếu xuất kho - tự động trừ tồn kho."""
    check_warehouse_access(current_user, data.kho_id)
    # Verify item exists
    item = db.query(Item).filter(Item.id == data.item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Không tìm thấy mặt hàng")
    if item.kho_id != data.kho_id:
        raise HTTPException(status_code=400, detail="Mặt hàng không thuộc kho này")

    # Check stock
    if item.ton_cuoi < data.so_luong:
        raise HTTPException(
            status_code=400,
            detail=f"Không đủ tồn kho. Tồn hiện tại: {item.ton_cuoi}, Yêu cầu: {data.so_luong}"
        )

    # Create transaction
    transaction = Transaction(
        loai="XUAT",
        item_id=item.id,
        ngay=data.ngay,
        ma_quan_ly=item.ma_quan_ly,
        ten_hang=item.ten_hang,
        ma_so=item.ma_so,
        so_luong=data.so_luong,
        don_vi_tinh=item.don_vi_tinh,
        cong_doan=data.cong_doan,
        nguoi_yeu_cau=data.nguoi_yeu_cau,
        nguoi_nhan=data.nguoi_nhan,
        nguoi_xuat=data.nguoi_xuat,
        trang_thai=data.trang_thai,
        ghi_chu=data.ghi_chu,
        kho_id=data.kho_id or item.kho_id
    )
    db.add(transaction)

    # Update item stock
    item.tong_xuat += data.so_luong
    item.ton_cuoi = item.ton_dau + item.tong_nhap - item.tong_xuat

    db.commit()
    db.refresh(transaction)
    return transaction





@router.get("/export-excel")
def export_transactions_excel(
    loai: Optional[str] = None,
    search: Optional[str] = None,
    from_date: Optional[date] = None,
    to_date: Optional[date] = None,
    cong_doan: Optional[str] = None,
    nguoi: Optional[str] = None,
    db: Session = Depends(get_db),
    _=Depends(require_permission("perm_excel"))
):
    """Xuất lịch sử giao dịch ra Excel."""
    from fastapi.responses import StreamingResponse
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Lich_Su_Giao_Dich"

    headers = ["STT", "Loại", "Ngày", "Mã QL", "Tên hàng", "Mã số", "Số lượng", "ĐVT", "Công đoạn", "Người thực hiện", "Ghi chú"]

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    query = db.query(Transaction)

    if loai:
        query = query.filter(Transaction.loai == loai)
    if search:
        search_lower = search.lower()
        query = query.filter(
            or_(
                func.lower(Transaction.ten_hang).contains(search_lower),
                func.lower(Transaction.ma_so).contains(search_lower),
            )
        )
    if from_date:
        query = query.filter(Transaction.ngay >= from_date)
    if to_date:
        query = query.filter(Transaction.ngay <= to_date)
    if cong_doan:
        query = query.filter(func.lower(Transaction.cong_doan).contains(cong_doan.lower()))
    if nguoi:
        nguoi_lower = nguoi.lower()
        query = query.filter(
            or_(
                func.lower(Transaction.nguoi_nhap).contains(nguoi_lower),
                func.lower(Transaction.nguoi_yeu_cau).contains(nguoi_lower),
            )
        )

    query = query.order_by(Transaction.id.desc())
    transactions = query.all()

    for row_idx, txn in enumerate(transactions, 2):
        nguoi_thuc_hien = txn.nguoi_nhap if txn.loai == "NHAP" else txn.nguoi_yeu_cau
        row_data = [
            row_idx - 1, txn.loai, txn.ngay.strftime("%Y-%m-%d") if txn.ngay else "", txn.ma_quan_ly,
            txn.ten_hang, txn.ma_so, txn.so_luong, txn.don_vi_tinh,
            txn.cong_doan, nguoi_thuc_hien, txn.ghi_chu
        ]
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border

    for col in ws.columns:
        max_length = 0
        for cell in col:
            if cell.value:
                max_length = min(max_length, len(str(cell.value))) # correction note: logic was adjusted in original snippet provided
        # simplified length calculation logic
        ws.column_dimensions[col[0].column_letter].width = 20

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    filename = "lich_su_nhap.xlsx" if loai == "NHAP" else "lich_su_xuat.xlsx" if loai == "XUAT" else "lich_su_giao_dich.xlsx"

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.delete("/{transaction_id}")
def delete_transaction(transaction_id: int, db: Session = Depends(get_db), _=Depends(require_permission("perm_delete"))):
    """Xoá giao dịch và hoàn lại tồn kho."""
    txn = db.query(Transaction).filter(Transaction.id == transaction_id).first()
    if not txn:
        raise HTTPException(status_code=404, detail="Không tìm thấy giao dịch")

    item = db.query(Item).filter(Item.id == txn.item_id).first()
    if item:
        if txn.loai == "NHAP":
            item.tong_nhap -= txn.so_luong
        else:
            item.tong_xuat -= txn.so_luong
        item.ton_cuoi = item.ton_dau + item.tong_nhap - item.tong_xuat

    db.delete(txn)
    db.commit()
    return {"message": "Đã xoá giao dịch và hoàn lại tồn kho"}
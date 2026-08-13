"""
API endpoints for Reports & Dashboard.
"""
from fastapi import APIRouter, Depends, Query
from sqlalchemy.orm import Session
from sqlalchemy import func, extract, case, or_
from typing import Optional
from datetime import date, timedelta

from ..deps import require_permission, apply_warehouse_filter, check_warehouse_access
from ..database import get_db
from ..models import Item, Transaction
from ..schemas import (
    DashboardStats, DashboardResponse, RecentTransaction,
    MonthlyTrend, TopItem, ItemResponse,
)

router = APIRouter(prefix="/api/reports", tags=["Báo cáo"])


@router.get("/dashboard", response_model=DashboardResponse)
def get_dashboard(kho_id: Optional[int] = None, db: Session = Depends(get_db), current_user = Depends(require_permission("perm_view"))):
    """Lấy dữ liệu dashboard."""
    today = date.today()

    # Stats
    item_query = db.query(Item)
    txn_query = db.query(Transaction)
    
    item_query = apply_warehouse_filter(item_query, Item, current_user, kho_id)
    txn_query = apply_warehouse_filter(txn_query, Transaction, current_user, kho_id)

    total_items = item_query.with_entities(func.count(Item.id)).scalar() or 0
    total_imports_today = (
        txn_query.with_entities(func.coalesce(func.sum(Transaction.so_luong), 0))
        .filter(Transaction.loai == "NHAP", Transaction.ngay == today)
        .scalar()
    )
    total_exports_today = (
        txn_query.with_entities(func.coalesce(func.sum(Transaction.so_luong), 0))
        .filter(Transaction.loai == "XUAT", Transaction.ngay == today)
        .scalar()
    )
    from sqlalchemy import or_
    low_stock_count = (
        item_query.with_entities(func.count(Item.id))
        .filter(or_(Item.ton_cuoi <= 0, Item.ton_cuoi <= Item.dinh_muc))
        .scalar() or 0
    )
    out_of_stock_count = (
        item_query.with_entities(func.count(Item.id))
        .filter(Item.ton_cuoi <= 0)
        .scalar() or 0
    )
    total_value = (
        item_query.with_entities(func.coalesce(func.sum(Item.don_gia * Item.ton_cuoi), 0))
        .scalar()
    )

    stats = DashboardStats(
        total_items=total_items,
        total_imports_today=total_imports_today,
        total_exports_today=total_exports_today,
        low_stock_count=low_stock_count,
        out_of_stock_count=out_of_stock_count,
        total_value=total_value,
    )

    # Recent transactions (last 10)
    recent_txns = (
        txn_query
        .order_by(Transaction.ngay.desc(), Transaction.id.desc())
        .limit(10)
        .all()
    )
    recent = []
    for t in recent_txns:
        nguoi = t.nguoi_nhap if t.loai == "NHAP" else t.nguoi_xuat
        recent.append(RecentTransaction(
            id=t.id,
            loai=t.loai,
            ten_hang=t.ten_hang,
            ma_so=t.ma_so,
            so_luong=t.so_luong,
            ngay=t.ngay,
            nguoi=nguoi,
        ))

    # Monthly trends (last 6 months)
    six_months_ago = today - timedelta(days=180)
    monthly_data_query = txn_query.with_entities(
            func.strftime('%Y-%m', Transaction.ngay).label('month'),
            Transaction.loai,
            func.sum(Transaction.so_luong).label('total'),
        )
    monthly_data = (
        monthly_data_query
        .filter(Transaction.ngay >= six_months_ago)
        .group_by(func.strftime('%Y-%m', Transaction.ngay), Transaction.loai)
        .all()
    )
    monthly_map = {}
    for row in monthly_data:
        if row.month not in monthly_map:
            monthly_map[row.month] = {"nhap": 0, "xuat": 0}
        if row.loai == "NHAP":
            monthly_map[row.month]["nhap"] = row.total
        else:
            monthly_map[row.month]["xuat"] = row.total

    monthly_trends = [
        MonthlyTrend(month=m, nhap=v["nhap"], xuat=v["xuat"])
        for m, v in sorted(monthly_map.items())
    ]

    # Top 10 exported items (last 30 days)
    thirty_days_ago = today - timedelta(days=30)
    top_exports_data = (
        txn_query.with_entities(
            Transaction.ten_hang,
            Transaction.ma_so,
            func.sum(Transaction.so_luong).label('total'),
        )
        .filter(Transaction.loai == "XUAT", Transaction.ngay >= thirty_days_ago)
        .group_by(Transaction.ten_hang, Transaction.ma_so)
        .order_by(func.sum(Transaction.so_luong).desc())
        .limit(10)
        .all()
    )
    top_exports = [
        TopItem(ten_hang=r.ten_hang, ma_so=r.ma_so, so_luong=r.total)
        for r in top_exports_data
    ]

    # Low stock items
    low_stock_items = (
        item_query
        .filter(or_(Item.ton_cuoi <= 0, Item.ton_cuoi <= Item.dinh_muc))
        .order_by(Item.ton_cuoi)
        .limit(20)
        .all()
    )

    return DashboardResponse(
        stats=stats,
        recent_transactions=recent,
        monthly_trends=monthly_trends,
        top_exports=top_exports,
        low_stock_items=[ItemResponse.model_validate(i) for i in low_stock_items],
    )


@router.get("/export/inventory")
def export_inventory_excel(
    kho_id: Optional[int] = None,
    search: Optional[str] = None,
    ma_quan_ly: Optional[str] = None,
    trang_thai: Optional[str] = None,
    cong_doan: Optional[str] = None,
    sort_by: Optional[str] = "id",
    sort_dir: Optional[str] = "asc",
    low_stock: Optional[bool] = None,
    loai_vat_tu: Optional[str] = None,
    db: Session = Depends(get_db), 
    current_user = Depends(require_permission("perm_excel"))
):
    """Xuất toàn bộ tồn kho ra Excel."""
    from fastapi.responses import StreamingResponse
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()

    # ---- Sheet 1: Tồn kho ----
    ws = wb.active
    ws.title = "TON_KHO"

    headers = ["STT", "Mã QL", "Tên hàng", "Mã số", "NCC", "Đơn giá", "Vị trí",
               "ĐVT", "Tồn đầu", "Nhập", "Xuất", "Tồn cuối", "Định mức", "Trạng thái", "Công đoạn", "Ghi chú"]

    # Header style
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # Data query
    query = db.query(Item)
    query = apply_warehouse_filter(query, Item, current_user, kho_id)

    # Search
    if search:
        search_lower = search.lower()
        query = query.filter(
            or_(
                func.lower(Item.ten_hang).contains(search_lower),
                func.lower(Item.ma_so).contains(search_lower),
                func.lower(Item.vi_tri).contains(search_lower),
                func.lower(Item.ma_quan_ly).contains(search_lower),
            )
        )

    # Filters
    if ma_quan_ly:
        query = query.filter(Item.ma_quan_ly == ma_quan_ly)
    if trang_thai:
        query = query.filter(Item.trang_thai == trang_thai)
    if cong_doan:
        query = query.filter(func.lower(Item.cong_doan).contains(cong_doan.lower()))
    if loai_vat_tu:
        query = query.filter(Item.loai_vat_tu == loai_vat_tu)
    if low_stock:
        query = query.filter(or_(Item.ton_cuoi <= 0, Item.ton_cuoi <= Item.dinh_muc))

    # Sort
    sort_column = getattr(Item, sort_by, Item.id)
    if sort_dir == "desc":
        sort_column = sort_column.desc()
    query = query.order_by(sort_column)

    items = query.all()
    for row_idx, item in enumerate(items, 2):
        row_data = [
            row_idx - 1, item.ma_quan_ly, item.ten_hang, item.ma_so,
            item.nha_cung_cap, item.don_gia, item.vi_tri, item.don_vi_tinh,
            item.ton_dau, item.tong_nhap, item.tong_xuat, item.ton_cuoi,
            item.dinh_muc, item.trang_thai, item.cong_doan, item.ghi_chu,
        ]
        
        is_low_stock = False
        try:
            if item.ton_cuoi == 0 or (item.dinh_muc and item.ton_cuoi <= item.dinh_muc):
                is_low_stock = True
        except:
            pass
            
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if col_idx == 12 and is_low_stock: # Column 12 is ton_cuoi
                cell.font = Font(color="FF0000", bold=True)

    # Auto-fit column widths
    for col in ws.columns:
        max_length = 0
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 40)

    # Save to stream
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=ton_kho.xlsx"},
    )

from pydantic import BaseModel
from typing import List
from ..mail_utils import send_warning_email

class SendEmailRequest(BaseModel):
    emails: List[str]
    kho_id: Optional[int] = None

@router.post("/send-warning-email")
def api_send_warning_email(
    request: SendEmailRequest,
    db: Session = Depends(get_db),
    current_user = Depends(require_permission("perm_excel"))
):
    try:
        # Fetch low stock items (tồn cuối <= định mức, bao gồm hàng hết trong kho)
        query = db.query(Item).filter(
            Item.ton_cuoi <= Item.dinh_muc
        )
        query = apply_warehouse_filter(query, Item, current_user, request.kho_id)
        items = query.all()
        
        # Lấy thông tin kho để hiển thị tên
        from ..models import Warehouse
        warehouse_name = "Tổng hợp"
        if request.kho_id:
            wh = db.query(Warehouse).filter(Warehouse.id == request.kho_id).first()
            if wh:
                warehouse_name = wh.ten_kho
                
        send_warning_email(request.emails, items, warehouse_name)
        return {"message": f"Đã gửi email thành công tới {len(request.emails)} địa chỉ."}
    except Exception as e:
        from fastapi import HTTPException
        raise HTTPException(status_code=500, detail=str(e))

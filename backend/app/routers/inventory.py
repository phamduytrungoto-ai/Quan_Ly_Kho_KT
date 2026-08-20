"""
API endpoints for Inventory (Tá»“n kho) management.
"""
import math
import os
import shutil
import json
import uuid
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from sqlalchemy.orm import Session
from typing import Optional
from datetime import date
from sqlalchemy import or_, func, literal_column

from .. import deps
from ..deps import apply_warehouse_filter, check_warehouse_access, check_warehouse_permission
from ..database import get_db
from ..logger import log_action
from ..models import Item, Transaction
from ..schemas import (
    ItemCreate, ItemUpdate, ItemResponse,
    ItemListResponse,
)

router = APIRouter(prefix="/api/items", tags=["Tá»“n kho"])


@router.get("", response_model=ItemListResponse)
def list_items(
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
    search: Optional[str] = None,
    ma_quan_ly: Optional[str] = None,
    trang_thai: Optional[str] = None,
    cong_doan: Optional[str] = None,
    sort_by: Optional[str] = "id",
    sort_dir: Optional[str] = "asc",
    low_stock: Optional[bool] = None,
    from_date: Optional[date] = None,
    to_date: Optional[date] = None,
    kho_id: Optional[int] = None,
    loai_vat_tu: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user = Depends(deps.get_current_user)
):
    """Lấy danh sách tồn kho với phân trang, tìm kiếm, lọc."""
    query = db.query(Item)
    query = apply_warehouse_filter(query, Item, current_user, db, kho_id)

    # Search
    if search:
        search_lower = search.lower()
        query = query.filter(
            or_(
                func.lower(Item.ten_hang).contains(search_lower),
                func.lower(Item.ma_so).contains(search_lower),
                func.lower(Item.vi_tri).contains(search_lower),
                func.lower(Item.ma_quan_ly).contains(search_lower),
            )
        )

    # Filters
    if loai_vat_tu:
        query = query.filter(Item.loai_vat_tu == loai_vat_tu)
        
    if ma_quan_ly:
        query = query.filter(Item.ma_quan_ly == ma_quan_ly)
    if trang_thai:
        query = query.filter(Item.trang_thai == trang_thai)
    if cong_doan:
        query = query.filter(func.lower(Item.cong_doan).contains(cong_doan.lower()))
    if low_stock:
        query = query.filter(Item.ton_cuoi <= Item.dinh_muc)

    # Count
    total = query.count()

    # Sort
    sort_column = getattr(Item, sort_by, Item.id)
    if sort_dir == "desc":
        sort_column = sort_column.desc()
    query = query.order_by(sort_column)

    # Paginate
    offset = (page - 1) * page_size
    items = query.offset(offset).limit(page_size).all()

    # Dynamic calculation if dates are provided
    results = []
    if from_date or to_date:
        item_ids = [item.id for item in items]
        if item_ids:
            # Query transactions for these items
            tx_query = db.query(
                Transaction.item_id,
                Transaction.loai,
                func.sum(Transaction.so_luong).label('total'),
                (Transaction.ngay < from_date).label('is_before') if from_date else literal_column("0").label('is_before')
            ).filter(Transaction.item_id.in_(item_ids))

            if to_date:
                tx_query = tx_query.filter(Transaction.ngay <= to_date)
            
            from sqlalchemy import literal_column
            tx_query = tx_query.group_by(
                Transaction.item_id, 
                Transaction.loai,
                (Transaction.ngay < from_date) if from_date else literal_column("0")
            ).all()

            # Group transactions
            tx_map = {}
            for row in tx_query:
                i_id, loai, total, is_before = row
                if i_id not in tx_map:
                    tx_map[i_id] = {'nhap_before': 0, 'xuat_before': 0, 'nhap_in': 0, 'xuat_in': 0}
                if is_before:
                    if loai == 'NHAP': tx_map[i_id]['nhap_before'] += total
                    elif loai == 'XUAT': tx_map[i_id]['xuat_before'] += total
                else:
                    if loai == 'NHAP': tx_map[i_id]['nhap_in'] += total
                    elif loai == 'XUAT': tx_map[i_id]['xuat_in'] += total

            for item in items:
                stats = tx_map.get(item.id, {'nhap_before': 0, 'xuat_before': 0, 'nhap_in': 0, 'xuat_in': 0})
                
                dyn_ton_dau = item.ton_dau + stats['nhap_before'] - stats['xuat_before']
                dyn_nhap = stats['nhap_in']
                dyn_xuat = stats['xuat_in']
                dyn_ton_cuoi = dyn_ton_dau + dyn_nhap - dyn_xuat

                # Create response object and override values
                resp = ItemResponse.model_validate(item)
                resp.ton_dau = dyn_ton_dau
                resp.tong_nhap = dyn_nhap
                resp.tong_xuat = dyn_xuat
                resp.ton_cuoi = dyn_ton_cuoi
                results.append(resp)
        else:
            results = [ItemResponse.model_validate(item) for item in items]
    else:
        results = [ItemResponse.model_validate(item) for item in items]

    return ItemListResponse(
        items=results,
        total=total,
        page=page,
        page_size=page_size,
        total_pages=math.ceil(total / page_size) if total > 0 else 1,
    )


@router.get("/all")
def list_all_items(
    search: Optional[str] = None,
    kho_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user = Depends(deps.get_current_user)
):
    """Lấy tất cả items (cho autocomplete). Trả về id, ten_hang, ma_so, don_vi_tinh, ton_cuoi."""
    query = db.query(Item.id, Item.ten_hang, Item.ma_so, Item.don_vi_tinh, Item.ton_cuoi, Item.ma_quan_ly)
    
    query = apply_warehouse_filter(query, Item, current_user, db, kho_id)
    
    if search:
        search_lower = search.lower()
        query = query.filter(
            or_(
                func.lower(Item.ten_hang).contains(search_lower),
                func.lower(Item.ma_so).contains(search_lower),
            )
        )
    items = query.order_by(Item.ten_hang).all()
    return [
        {
            "id": i.id,
            "ten_hang": i.ten_hang,
            "ma_so": i.ma_so,
            "don_vi_tinh": i.don_vi_tinh,
            "ton_cuoi": i.ton_cuoi,
            "ma_quan_ly": i.ma_quan_ly,
        }
        for i in items
    ]


@router.get("/{item_id}", response_model=ItemResponse)
def get_item(item_id: int, db: Session = Depends(get_db), current_user = Depends(deps.get_current_user)):
    item = db.query(Item).filter(Item.id == item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Khong tim thay Item")
    check_warehouse_permission(current_user, item.kho_id, "perm_view", db)
    return item


@router.post("", response_model=ItemResponse, status_code=201)
def create_item(data: ItemCreate, db: Session = Depends(get_db), current_user = Depends(deps.get_current_user)):
    check_warehouse_permission(current_user, data.kho_id, "perm_add", db)
    # Check if duplicate in the same warehouse
    kho_id = data.kho_id if data.kho_id else 1
    existing = db.query(Item).filter(
        Item.ma_so == data.ma_so,
        Item.kho_id == kho_id
    ).first()
    if existing:
        raise HTTPException(
            status_code=400,
            detail=f"Mặt hàng với mã số {data.ma_so} đã tồn tại trong kho này."
        )
    new_item = Item(**data.model_dump())
    new_item.ton_cuoi = new_item.ton_dau
    db.add(new_item)
    db.commit()
    db.refresh(new_item)
    log_action(db, None, current_user, "Thêm hàng hóa", f"Mã quản lý: {new_item.ma_quan_ly}, Mã số: {new_item.ma_so}, Tên: {new_item.ten_hang}")
    return new_item


@router.put("/{item_id}", response_model=ItemResponse)
def update_item(item_id: int, data: ItemUpdate, db: Session = Depends(get_db), current_user = Depends(deps.get_current_user)):
    item = db.query(Item).filter(Item.id == item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Item not found")
    check_warehouse_permission(current_user, item.kho_id, "perm_edit", db)

    if data.ma_so:
        existing = db.query(Item).filter(
            Item.ma_so == data.ma_so, 
            Item.kho_id == item.kho_id,
            Item.id != item_id
        ).first()
        if existing:
            raise HTTPException(
                status_code=400,
                detail=f"Mã số {data.ma_so} đã được sử dụng cho một mặt hàng khác trong kho này."
            )

    update_data = data.model_dump(exclude_unset=True)
    for key, value in update_data.items():
        setattr(item, key, value)

    db.commit()
    db.refresh(item)
    log_action(db, None, current_user, "Cập nhật hàng hóa", f"Mã số: {item.ma_so}, Tên: {item.ten_hang}")
    return item


@router.delete("/{item_id}")
def delete_item(
    item_id: int, 
    db: Session = Depends(get_db),
    current_user = Depends(deps.get_current_user)
):
    item = db.query(Item).filter(Item.id == item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Item not found")
    check_warehouse_permission(current_user, item.kho_id, "perm_delete", db)

    # Check if there are transactions
    has_transactions = db.query(Transaction).filter(Transaction.item_id == item_id).first() is not None
    if has_transactions:
        raise HTTPException(status_code=400, detail="Cannot delete item with transaction history")

    db.delete(item)
    db.commit()
    log_action(db, None, current_user, "Xóa hàng hóa", f"Mã số: {item.ma_so}, Tên: {item.ten_hang}")
    return {"message": "Deleted successfully"}

@router.post("/{item_id}/image")
def upload_item_image(
    item_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user = Depends(deps.get_current_user)
):
    item = db.query(Item).filter(Item.id == item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Item not found")
    check_warehouse_permission(current_user, item.kho_id, "perm_edit", db)

    # Validate file type
    allowed_extensions = {".jpg", ".jpeg", ".png", ".gif", ".webp"}
    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in allowed_extensions:
        raise HTTPException(status_code=400, detail="Invalid image format. Allowed: JPG, PNG, GIF, WEBP")

    # Create uploads directory if not exists
    upload_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "..", "uploads", "items")
    os.makedirs(upload_dir, exist_ok=True)

    # Generate unique filename
    unique_filename = f"{uuid.uuid4().hex}{ext}"
    file_path = os.path.join(upload_dir, unique_filename)

    # Save file
    try:
        with open(file_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Could not save file: {str(e)}")

    # Update item model to support multiple images
    old_image_data = item.hinh_anh
    images = []
    if old_image_data:
        try:
            images = json.loads(old_image_data)
            if not isinstance(images, list):
                images = [old_image_data] # Fallback for legacy single string
        except:
            images = [old_image_data] # Fallback for legacy single string
    
    new_url = f"/uploads/items/{unique_filename}"
    images.append(new_url)
    item.hinh_anh = json.dumps(images)
    db.commit()

    return {"message": "Image uploaded successfully", "url": new_url, "images": images}

@router.get("/import/template")
def download_import_template(current_user = Depends(deps.get_current_user)):
    """Tải file Excel mẫu để nhập tồn kho."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from fastapi.responses import StreamingResponse
    import io

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Nhập tồn kho"

    # Headers
    headers = [
        ("Mã quản lý", 15),
        ("Tên hàng", 40),
        ("Mã số", 20),
        ("Nhà cung cấp", 25),
        ("Đơn giá", 15),
        ("Vị trí", 15),
        ("ĐVT", 10),
        ("Tồn kho", 12),
        ("Định mức", 12),
        ("Công đoạn", 20),
        ("Loại vật tư", 20),
        ("Thông số kỹ thuật", 30),
        ("Ghi chú", 25),
    ]

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    for col_idx, (name, width) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    # Sample row
    sample = ["BE", "BEARING 6200", "6200-2RS", "NSK", 50000, "Kệ A1", "PCS", 10, 5, "Sản xuất", "Vật tư dự phòng", "20x30x9mm", "Hàng mới"]
    for col_idx, val in enumerate(sample, 1):
        cell = ws.cell(row=2, column=col_idx, value=val)
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center")

    # Note row
    note_cell = ws.cell(row=4, column=1, value="Lưu ý: Cột 'Mã số' là bắt buộc. Các cột khác có thể để trống.")
    note_cell.font = Font(italic=True, color="FF6600")
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=6)

    note2 = ws.cell(row=5, column=1, value="Loại vật tư: 'Vật tư tiêu hao', 'Vật tư dự phòng', hoặc 'Công cụ dụng cụ'")
    note2.font = Font(italic=True, color="888888")
    ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=6)

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=Mau_Nhap_Ton_Kho.xlsx"}
    )


@router.post("/import")
async def import_excel(
    file: UploadFile = File(...),
    kho_id: int = Query(1),
    update_existing: bool = Query(False),
    db: Session = Depends(get_db),
    current_user = Depends(deps.get_current_user)
):
    check_warehouse_permission(current_user, kho_id, "perm_add", db)
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Chỉ hỗ trợ file Excel (.xlsx, .xls)")
        
    import pandas as pd
    import io
    
    contents = await file.read()
    try:
        df = pd.read_excel(io.BytesIO(contents))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Lỗi đọc file Excel: {str(e)}")

    def clean(val, default=""):
        """Chuyển giá trị từ pandas sang string, xử lý NaN."""
        s = str(val).strip() if val is not None else ""
        return default if s == "nan" or s == "" else s

    def clean_int(val, default=0):
        try:
            return int(float(val))
        except:
            return default

    def clean_float(val, default=0.0):
        try:
            return float(val)
        except:
            return default

    success_count = 0
    update_count = 0
    skip_count = 0
    errors = []
    
    for index, row in df.iterrows():
        row_num = index + 2  # Excel row number (1-indexed header + data)
        
        ma_so = clean(row.get('Mã số', ''))
        if not ma_so:
            ma_so = clean(row.get('Part No.', ''))
        if not ma_so:
            continue
            
        ten_hang = clean(row.get('Tên hàng', ''))
        
        # Check existing
        existing = db.query(Item).filter(Item.ma_so == ma_so, Item.kho_id == kho_id).first()
        
        if existing:
            if update_existing:
                # Cập nhật thông tin (KHÔNG thay đổi tồn kho)
                if ten_hang: existing.ten_hang = ten_hang
                ncc = clean(row.get('Nhà cung cấp', ''))
                if ncc: existing.nha_cung_cap = ncc
                don_gia = row.get('Đơn giá', None)
                if don_gia is not None and clean(don_gia): existing.don_gia = clean_float(don_gia)
                vi_tri = clean(row.get('Vị trí', ''))
                if vi_tri: existing.vi_tri = vi_tri
                dvt = clean(row.get('ĐVT', ''))
                if dvt: existing.don_vi_tinh = dvt
                dinh_muc = row.get('Định mức', None)
                if dinh_muc is not None and clean(dinh_muc): existing.dinh_muc = clean_int(dinh_muc)
                cong_doan = clean(row.get('Công đoạn', ''))
                if cong_doan: existing.cong_doan = cong_doan
                loai_vt = clean(row.get('Loại vật tư', ''))
                if loai_vt: existing.loai_vat_tu = loai_vt
                tskt = clean(row.get('Thông số kỹ thuật', ''))
                if tskt: existing.thong_so_ky_thuat = tskt
                ghi_chu = clean(row.get('Ghi chú', ''))
                if ghi_chu: existing.ghi_chu = ghi_chu
                ma_ql = clean(row.get('Mã quản lý', ''))
                if ma_ql: existing.ma_quan_ly = ma_ql
                update_count += 1
            else:
                skip_count += 1
            continue
        
        # Tạo mới
        ton_dau = clean_int(row.get('Tồn kho', 0))
            
        new_item = Item(
            ma_so=ma_so,
            ten_hang=ten_hang,
            kho_id=kho_id,
            ton_dau=ton_dau,
            ton_cuoi=ton_dau,
            ma_quan_ly=clean(row.get('Mã quản lý', '')),
            don_vi_tinh=clean(row.get('ĐVT', ''), 'PCS'),
            vi_tri=clean(row.get('Vị trí', '')),
            ghi_chu=clean(row.get('Ghi chú', '')),
            nha_cung_cap=clean(row.get('Nhà cung cấp', '')),
            don_gia=clean_float(row.get('Đơn giá', 0)),
            dinh_muc=clean_int(row.get('Định mức', 0)),
            cong_doan=clean(row.get('Công đoạn', '')),
            loai_vat_tu=clean(row.get('Loại vật tư', ''), 'Vật tư tiêu hao'),
            thong_so_ky_thuat=clean(row.get('Thông số kỹ thuật', '')),
        )
        db.add(new_item)
        success_count += 1
        
    db.commit()
    
    detail_parts = []
    if success_count: detail_parts.append(f"Thêm mới: {success_count}")
    if update_count: detail_parts.append(f"Cập nhật: {update_count}")
    if skip_count: detail_parts.append(f"Bỏ qua (đã tồn tại): {skip_count}")
    detail_msg = ", ".join(detail_parts) if detail_parts else "Không có dữ liệu"
    
    log_action(db, None, current_user, "Nhập tồn kho từ Excel", f"Kho ID: {kho_id}, {detail_msg}")
    
    return {
        "message": f"Nhập Excel hoàn tất. {detail_msg}",
        "success_count": success_count,
        "update_count": update_count,
        "skip_count": skip_count
    }


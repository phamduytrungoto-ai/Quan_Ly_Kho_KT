import os
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

from .database import SessionLocal
from .models import Setting

def get_db_setting(db, key, env_fallback):
    setting = db.query(Setting).filter(Setting.key == key).first()
    if setting and setting.value is not None:
        return setting.value
    return os.getenv(key, env_fallback)

def get_smtp_configs():
    db = SessionLocal()
    try:
        configs = []
        # Primary (e.g. Internal SMTP)
        host1 = get_db_setting(db, "SMTP_HOST", "")
        port1_str = get_db_setting(db, "SMTP_PORT", "587")
        port1 = int(port1_str) if port1_str else 587
        user1 = get_db_setting(db, "SMTP_USER", "")
        pass1 = get_db_setting(db, "SMTP_PASSWORD", "")
        if host1:
            configs.append((host1, port1, user1, pass1))
            
        # Fallback (e.g. Gmail for outside LAN)
        host2 = get_db_setting(db, "SMTP_FALLBACK_HOST", "")
        port2_str = get_db_setting(db, "SMTP_FALLBACK_PORT", "465")
        port2 = int(port2_str) if port2_str else 465
        user2 = get_db_setting(db, "SMTP_FALLBACK_USER", "")
        pass2 = get_db_setting(db, "SMTP_FALLBACK_PASSWORD", "")
        if host2:
            configs.append((host2, port2, user2, pass2))
            
        return configs
    finally:
        db.close()

def send_with_fallback(msg, default_sender):
    configs = get_smtp_configs()
    if not configs:
        raise ValueError("Chưa cấu hình SMTP_HOST trong file .env")
        
    last_error = None
    for host, port, user, pwd in configs:
        try:
            # Set sender email for this config
            sender = user if user else default_sender
            # Update From header if it already exists, otherwise add it
            if 'From' in msg:
                msg.replace_header('From', f"Hệ thống quản lý kho KT <{sender}>")
            else:
                msg['From'] = f"Hệ thống quản lý kho KT <{sender}>"
                
            if port == 465:
                server = smtplib.SMTP_SSL(host, port, timeout=10)
            else:
                server = smtplib.SMTP(host, port, timeout=10)
                try:
                    server.starttls()
                except smtplib.SMTPNotSupportedError:
                    pass
                except Exception as e:
                    pass
            
            if user and pwd:
                server.login(user, pwd)
                
            server.send_message(msg)
            try:
                server.quit()
            except Exception:
                pass
            return True # Success
        except Exception as e:
            last_error = e
            print(f"Failed to send via {host}:{port} - {e}")
            
    raise Exception(f"Không thể gửi email qua bất kỳ máy chủ nào. Lỗi cuối: {str(last_error)}")

def send_warning_email(recipient_emails: list, items: list, warehouse_name: str = "Tổng hợp"):
    
    html_content = f"""
    <html>
    <head>
        <style>
            h2 {{ font-family: Arial, sans-serif; color: #333; }}
            p {{ font-family: Arial, sans-serif; color: #555; line-height: 1.5; }}
        </style>
    </head>
    <body>
        <h2>⚠️ Cảnh Báo Tồn Kho Dưới Định Mức - {warehouse_name}</h2>
        <p>Hệ thống tự động gửi cảnh báo các sản phẩm đang có số lượng tồn kho bằng hoặc dưới mức định mức an toàn tại kho <b>{warehouse_name}</b>.</p>
        <p>Vui lòng xem chi tiết danh sách trong file Excel đính kèm.</p>
        <br>
        <p><i>Email được tạo và gửi tự động từ Hệ thống Quản lý Kho. Vui lòng không trả lời email này.</i></p>
    </body>
    </html>
    """

    # Create the email message
    msg = MIMEMultipart()
    msg['To'] = ", ".join(recipient_emails)
    msg['Subject'] = f"Hệ thống quản lý kho - cảnh báo tồn kho [{warehouse_name}]"
    
    msg.attach(MIMEText(html_content, 'html'))
    
    # Create Excel attachment
    try:
        import io
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from email.mime.application import MIMEApplication
        
        if items:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Cảnh Báo Tồn Kho"
            
            # Styles
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            align_center = Alignment(horizontal="center", vertical="center")
            align_left = Alignment(horizontal="left", vertical="center")
            align_right = Alignment(horizontal="right", vertical="center")
            thin_border = Border(
                left=Side(style='thin', color='E0E0E0'),
                right=Side(style='thin', color='E0E0E0'),
                top=Side(style='thin', color='E0E0E0'),
                bottom=Side(style='thin', color='E0E0E0')
            )
            
            red_bold_font = Font(color="FF0000", bold=True)
            orange_bold_font = Font(color="FFA500", bold=True)
            
            # Headers
            headers = ["STT", "Mã số", "Tên hàng", "Vị trí", "Tồn cuối", "Định mức"]
            ws.append(headers)
            
            # Apply header styles
            for col_num in range(1, 7):
                cell = ws.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
                if col_num in [1]:
                    cell.alignment = align_center
                elif col_num in [5, 6]:
                    cell.alignment = align_right
                else:
                    cell.alignment = align_left
            
            # Data
            for idx, item in enumerate(items, start=1):
                row_data = [
                    idx,
                    item.ma_so or '',
                    item.ten_hang or '',
                    item.vi_tri or '',
                    item.ton_cuoi,
                    item.dinh_muc
                ]
                ws.append(row_data)
                
                row_num = idx + 1
                for col_num in range(1, 7):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.border = thin_border
                    
                    if col_num == 1:
                        cell.alignment = align_center
                    elif col_num in [2, 3, 4]:
                        cell.alignment = align_left
                    elif col_num == 5:
                        cell.alignment = align_right
                        cell.font = red_bold_font
                    elif col_num == 6:
                        cell.alignment = align_right
                        cell.font = orange_bold_font
                
            # Adjust column widths
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 40
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 12
            ws.column_dimensions['F'].width = 12
                
            # Save to BytesIO
            excel_stream = io.BytesIO()
            wb.save(excel_stream)
            excel_stream.seek(0)
            
            part = MIMEApplication(excel_stream.read(), Name="Canh_Bao_Ton_Kho.xlsx")
            part['Content-Disposition'] = 'attachment; filename="Canh_Bao_Ton_Kho.xlsx"'
            msg.attach(part)
    except Exception as e:
        print(f"Không thể tạo file đính kèm Excel: {e}")
    
    send_with_fallback(msg, "no-reply@sharp-world.com")

def send_otp_email(recipient_email: str, otp: str):

    html_content = f"""
    <html>
    <head>
        <style>
            h2 {{ font-family: Arial, sans-serif; color: #333; }}
            .otp-box {{ font-size: 24px; font-weight: bold; padding: 15px; background: #f2f2f2; display: inline-block; letter-spacing: 5px; border-radius: 5px; }}
        </style>
    </head>
    <body>
        <h2>Mã xác thực Khôi phục mật khẩu</h2>
        <p>Hệ thống nhận được yêu cầu khôi phục mật khẩu cho tài khoản của bạn.</p>
        <p>Vui lòng nhập mã OTP dưới đây để tiến hành đổi mật khẩu. Mã này có hiệu lực trong vòng 10 phút:</p>
        <div class="otp-box">{otp}</div>
        <p>Nếu bạn không yêu cầu đổi mật khẩu, vui lòng bỏ qua email này.</p>
        <br>
        <p><i>Hệ thống quản lý kho KT</i></p>
    </body>
    </html>
    """

    msg = MIMEMultipart()
    msg['To'] = recipient_email
    msg['Subject'] = "Hệ thống quản lý kho KT"
    
    msg.attach(MIMEText(html_content, 'html'))
    
    send_with_fallback(msg, "no-reply@sharp-world.com")

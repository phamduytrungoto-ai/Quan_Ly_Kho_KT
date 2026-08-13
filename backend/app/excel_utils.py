import io
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from datetime import datetime

def generate_receipt_excel(data: dict, receipt_type: str = "issue") -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Phiếu In"

    # Settings and variables
    ma_kho = data.get("ma_kho", "DP-EE")
    
    if receipt_type == 'issue':
        title1 = f"PHIẾU YÊU CẦU XUẤT KHO ({ma_kho})"
        title2 = f"倉庫出荷依頼書 ({ma_kho})"
    elif receipt_type == 'receipt':
        title1 = f"PHIẾU NHẬP KHO ({ma_kho})"
        title2 = f"倉庫入庫依頼書 ({ma_kho})"
    else:
        title1 = f"PHIẾU CHUYỂN KHO ({ma_kho})"
        title2 = f"倉庫移動依頼書 ({ma_kho})"
        
    date_val = data.get("ngay_xuat") or data.get("ngay_nhap") or data.get("ngay_chuyen")
    if date_val and isinstance(date_val, str):
        # Format string to DD/MM/YYYY if possible
        try:
            date_obj = datetime.strptime(date_val.split('T')[0], "%Y-%m-%d")
            date_str = date_obj.strftime("%d/%m/%Y")
        except:
            date_str = date_val
    elif date_val:
        date_str = date_val.strftime("%d/%m/%Y")
    else:
        date_str = ""
        
    nguoi_thuc_hien_str = 'Người yêu cầu / 要求者名前'
    nguoi_thuc_hien = data.get("nguoi_yeu_cau") or data.get("nguoi_nhap") or data.get("nguoi_chuyen") or ""
    msnv = ""
    if "-" in nguoi_thuc_hien:
        parts = nguoi_thuc_hien.split("-")
        msnv = parts.pop().strip()
        nguoi_thuc_hien = "-".join(parts).strip()

    loai_xuat = data.get("loai_xuat", "")
    is_cap_moi = loai_xuat in ["Cấp mới", "Cấp Mới"]
    is_thay_the = loai_xuat in ["Thay thế", "Thay Thế"]
    cap_moi_char = "V" if is_cap_moi else ""
    thay_the_char = "V" if is_thay_the else ""

    # Border definitions
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    # Set column widths
    col_widths = {
        'A': 5,  'B': 25, 'C': 15, 'D': 10, 'E': 8, 'F': 12,
        'G': 10, 'H': 10, 'I': 8,  'J': 8,  'K': 15
    }
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    # Row 1: Logo & Form No
    ws.merge_cells('A1:B1')
    ws['A1'] = "SHARP"
    ws['A1'].font = Font(bold=True, color="FF0000", size=14)
    ws['A1'].alignment = Alignment(vertical="center", horizontal="left")
    
    ws.merge_cells('C1:K1')
    ws['C1'] = "Form No: Q-FOBV-KPE-060 Revision: 000 Issue date: 13/11/2023"
    ws['C1'].font = Font(size=10)
    ws['C1'].alignment = Alignment(vertical="center", horizontal="right")
    
    for row in ws['A1:K1']:
        for cell in row:
            cell.border = thin_border

    # Row 2 & 3: Title
    ws.merge_cells('A2:K3')
    ws['A2'] = f"{title1}\n{title2}"
    ws['A2'].font = Font(bold=True, size=14)
    ws['A2'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws['A2:K3']:
        for cell in row:
            cell.border = thin_border
    ws.row_dimensions[2].height = 25
    ws.row_dimensions[3].height = 20

    # Row 4: Info (Ngày, Người yêu cầu)
    ws.merge_cells('A4:E4')
    ws['A4'] = f"Ngày / 日付 : {date_str}"
    ws.merge_cells('F4:K4')
    ws['F4'] = f"{nguoi_thuc_hien_str} : {nguoi_thuc_hien}"
    
    # Row 5: Info (Bộ phận, MSNV)
    ws.merge_cells('A5:E5')
    ws['A5'] = f"Bộ phận / 部門 : {ma_kho}"
    ws.merge_cells('F5:K5')
    ws['F5'] = f"MSNV / ID : {msnv}"
    
    for row in ws['A4:K5']:
        for cell in row:
            cell.border = thin_border
            cell.font = Font(size=11)
            cell.alignment = Alignment(vertical="center")

    # Row 6: Table Headers
    if receipt_type == 'transfer':
        header_cap_moi = "Vị trí (Từ)\n移動元"
        header_thay_the = "Vị trí (Đến)\n移動先"
    else:
        header_cap_moi = "Cấp mới\n新規提供"
        header_thay_the = "Thay thế\n交換"

    headers = [
        "STT\nNo.", "Tên linh kiện\n部品名", "Mã linh kiện\n部品番号", 
        "SL yêu cầu\n依頼数量", "Đơn vị\n単位", "Công đoạn sử dụng\n使用工程", 
        "SL xuất kho\n倉庫出荷数量", "SL còn lại\n残り数量", header_cap_moi, 
        header_thay_the, "Ghi chú\n備考"
    ]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col_idx, value=header)
        cell.font = Font(bold=True, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        cell.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    ws.row_dimensions[6].height = 30

    # Table Data
    transactions = data.get("transactions", [])
    row_count = max(5, len(transactions))
    current_row = 7
    
    for i in range(row_count):
        is_low_stock = False
        if i < len(transactions):
            tx = transactions[i]
            ton_cuoi_val = tx.get("ton_cuoi", 0)
            dinh_muc_val = tx.get("dinh_muc", 0)
            try:
                ton_cuoi_val = float(ton_cuoi_val)
                dinh_muc_val = float(dinh_muc_val)
                if ton_cuoi_val == 0 or ton_cuoi_val <= dinh_muc_val:
                    is_low_stock = True
            except:
                pass

            if receipt_type == 'transfer':
                val_cap_moi = tx.get("vi_tri_cu", "")
                val_thay_the = tx.get("vi_tri_moi", "")
            else:
                val_cap_moi = cap_moi_char
                val_thay_the = thay_the_char

            row_data = [
                i + 1,
                tx.get("ten_hang", ""),
                tx.get("ma_so", ""),
                tx.get("so_luong", ""),
                tx.get("don_vi_tinh", ""),
                tx.get("cong_doan", ""),
                tx.get("so_luong", ""),
                tx.get("ton_cuoi", ""),
                val_cap_moi,
                val_thay_the,
                tx.get("ghi_chu", "")
            ]
        else:
            row_data = ["", "", "", "", "", "", "", "", "", "", ""]
            
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=val)
            cell.border = thin_border
            if col_idx == 8 and is_low_stock:
                cell.font = Font(size=11, color="FF0000", bold=True)
            else:
                cell.font = Font(size=11)
                
            if col_idx in [1, 4, 5, 6, 7, 8, 9, 10]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(vertical="center")
        current_row += 1

    # Signature Row 1 (Titles)
    ws.merge_cells(f'A{current_row}:C{current_row}')
    ws.cell(row=current_row, column=1, value="Nhân viên kho (ký tên)\n倉庫者 (サイン)")
    
    ws.merge_cells(f'D{current_row}:G{current_row}')
    ws.cell(row=current_row, column=4, value="Quản lý kho (ký tên)\n倉庫管理者 (サイン)")
    
    ws.merge_cells(f'H{current_row}:K{current_row}')
    ws.cell(row=current_row, column=8, value="Người yêu cầu (ký tên)\n要求者 (サイン)")

    for col in [1, 4, 8]:
        cell = ws.cell(row=current_row, column=col)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.font = Font(size=11)

    for row in ws[f'A{current_row}:K{current_row}']:
        for cell in row:
            cell.border = thin_border
            
    ws.row_dimensions[current_row].height = 30
    current_row += 1

    # Signature Row 2 (Boxes)
    ws.merge_cells(f'A{current_row}:C{current_row}')
    ws.merge_cells(f'D{current_row}:G{current_row}')
    ws.merge_cells(f'H{current_row}:K{current_row}')

    for row in ws[f'A{current_row}:K{current_row}']:
        for cell in row:
            cell.border = thin_border
    ws.row_dimensions[current_row].height = 70
    current_row += 1

    # Footer
    ws.merge_cells(f'A{current_row}:K{current_row}')
    footer_cell = ws.cell(row=current_row, column=1, value="Sharp manufacturing Vietnam")
    footer_cell.font = Font(bold=True, size=11)
    footer_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
